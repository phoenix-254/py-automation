import openpyxl as xl
from openpyxl.chart import BarChart, Reference


class FileProcessor:
    @staticmethod
    def process_file(file_path):
        workbook = xl.load_workbook(file_path)
        sheet = workbook['Sheet1']

        FileProcessor.__correct_data(sheet)

        FileProcessor.__add_chart(sheet)

        workbook.save(file_path)

    @staticmethod
    def __correct_data(sheet):
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 3)
            cell.value = cell.value * 0.9

    @staticmethod
    def __add_chart(sheet):
        values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=3, max_col=3)

        chart = BarChart()
        chart.add_data(values)
        sheet.add_chart(chart, 'D2')
